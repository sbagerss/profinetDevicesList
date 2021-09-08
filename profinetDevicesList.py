import clr
clr.AddReference('C:\\Program Files\\Siemens\\Automation\\Portal V16\\PublicAPI\\V16\\Siemens.Engineering.dll')
import Siemens.Engineering as tia
import Siemens.Engineering.HW.Features as hwf
from openpyxl import Workbook
from openpyxl.styles import Font
import argparse

############################################################################
class DevIpInfo:
    pnDeviceName = ""
    address = ""

    ########################################################################
    def IsGreater(self, obj):
        numberList1 = DevIpInfo.__convertIpAddress(self.address)
        numberList2 = DevIpInfo.__convertIpAddress(obj.address)

        if len(numberList1) > len(numberList2):
            return True
        elif len(numberList1) < len(numberList2):
            return False
        else:
            i = 0
            while i < len(numberList1):
                if numberList1[i] > numberList2[i]:
                    return True
                elif numberList1[i] < numberList2[i]:
                    return False
                i = i + 1
            return False

    ########################################################################
    def __convertIpAddress(addressString):
        numberList = []
        for n in addressString.split("."):
            if n.isnumeric():
                numberList.append(int(n))
        return numberList

############################################################################
def showAllAttributes(device, prefix = ""):
    print(prefix + "showAllAttributes__________________________")
    try:
        for atr in device.GetAttributeInfos():
            print(prefix + atr.Name + ": " + str(device.GetAttribute(atr.Name)))
    except:
        print("BRAK")

############################################################################
def deviceItemsPresent(device):
    try:
        x = device.DeviceItems[0]
        return True
    except:
        return False

############################################################################
def getAllDeviceItems(device, devLista):
    devLista.append(device)
    if deviceItemsPresent(device):
        for d in device.DeviceItems:
            getAllDeviceItems(d, devLista)

############################################################################

parser = argparse.ArgumentParser()
parser.add_argument('filename', help="Nazwa pliku wynikowego")
args = parser.parse_args()

processes = tia.TiaPortal.GetProcesses() 
mytia = processes[0].Attach()
tiaProject = mytia.Projects[0]

devLista = []
for dev in tiaProject.Devices:
    getAllDeviceItems(dev, devLista)

for dev in tiaProject.UngroupedDevicesGroup.Devices:
    getAllDeviceItems(dev, devLista)

networkServicesList = []
for d in devLista:
    networkService = tia.IEngineeringServiceProvider(d).GetService[hwf.NetworkInterface]()
    if type(networkService) is hwf.NetworkInterface:
        networkServicesList.append(networkService)

ipList = []
for netService in networkServicesList:
    for n in netService.Nodes:
        connected = n.GetAttribute("ConnectedSubnet")
        if connected != None:
            obj = DevIpInfo()
            obj.pnDeviceName = n.GetAttribute("PnDeviceName")
            obj.address = n.GetAttribute("Address")
            ipList.append(obj)

changed = True
while changed:
    i = 0
    changed = False
    while i < len(ipList)-1:
        if(ipList[i].IsGreater(ipList[i+1])):
            obj = ipList[i+1]
            ipList[i+1] = ipList[i]
            ipList[i] = obj
            changed = True
        i = i + 1

wb = Workbook()
sheet = wb.active

ft = Font(bold=True, underline="single")
sheet.cell(row = 1, column = 1).value = "profinet name"
sheet.cell(row = 1, column = 2).value = "address"
sheet.cell(row = 1, column = 1).font = ft
sheet.cell(row = 1, column = 2).font = ft

i = 2
for n in ipList:
    sheet.cell(row = i, column = 1).value = "'" + n.pnDeviceName + "'"
    sheet.cell(row = i, column = 2).value = n.address
    i = i + 1

wb.save(str(args.filename) + '.xlsx')
